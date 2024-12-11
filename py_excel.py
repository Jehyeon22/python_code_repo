# 엑셀 모듈 불러오기
from openpyxl import Workbook

# 새로운 엑셀 문서 생성
write_wb = Workbook()

# 이름이 있는 워크시트를 생성
write_ws = write_wb.create_sheet('테스트용 시트')

# 시트를 활성화
write_ws = write_wb['테스트용 시트']

# 데이터 입력
write_ws['A1'] = '숫자'
write_ws.append([1, 2, 3])
write_ws.cell(5, 5, '5행5열')

# 사용 가능한 시트 이름 출력 (확인용)
print(write_wb.sheetnames)

# 엑셀 파일 저장 (파일 경로 확인 및 권한 문제 해결)
write_wb.save(r"C:\study for python\python_excel\숫자.xlsx")
